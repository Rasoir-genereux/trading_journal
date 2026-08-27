"""
Parses a MetaTrader 5 "Report History" .xlsx export into closed trades.

Unlike TradingView, whose export lists individual orders that must be
FIFO-matched into trades, MT5's "Positions" section already gives one row
per closed position - partial closes are pre-aggregated by the platform
itself - so no reconstruction is needed here, just column mapping.

The "Profit" column in that section is pure price P&L, already converted to
the account's deposit currency; commission and swap are reported separately
and are combined here into the trade's `commission` field, matching how the
rest of the app treats commission as a cost tracked apart from P&L (and
subtracted from it for display). MT5 reports commission/swap as negative
numbers (a debit); the app's convention is the opposite - a positive
`commission` value is a cost - so the sign is flipped here.
"""
import io
import re
import unicodedata
from datetime import datetime

import openpyxl

from csv_import import clean_symbol

# MT5 generates this report in whatever language the terminal is set to; column
# labels differ (e.g. "Time"/"Heure", "Symbol"/"Symbole", "Swap"/"Echange") but the
# underlying report structure doesn't. Match headers against known aliases in each
# language rather than assuming English, so a French (or other) terminal still works.
HEADER_ALIASES = {
    "time": {"time", "heure"},
    "position": {"position"},
    "symbol": {"symbol", "symbole"},
    "type": {"type"},
    "volume": {"volume"},
    "price": {"price", "prix"},
    "commission": {"commission"},
    "swap": {"swap", "echange"},
    "profit": {"profit"},
}
REQUIRED_CANONICAL = {"position", "symbol", "type", "volume", "commission", "swap", "profit"}
ACCOUNT_LABELS = {"account", "compte"}


class Mt5FormatError(Exception):
    pass


def _norm(h):
    s = str(h or "").strip().lower()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"[\s/]+", "", s)


def _canonical_header(cell):
    normed = _norm(cell)
    for canon, aliases in HEADER_ALIASES.items():
        if normed in aliases:
            return canon
    return None


def _account_currency(ws):
    for row in ws.iter_rows(max_row=15, values_only=True):
        if row and isinstance(row[0], str) and row[0].strip().lower().rstrip(":") in ACCOUNT_LABELS:
            for cell in row[1:]:
                if isinstance(cell, str):
                    m = re.search(r"\(([A-Za-z]{3})\s*,", cell)
                    if m:
                        return m.group(1).upper()
    return None


def _parse_time(value):
    if not value:
        return None
    try:
        return datetime.strptime(str(value).strip(), "%Y.%m.%d %H:%M:%S").isoformat()
    except ValueError:
        return None


def parse_xlsx(file_bytes: bytes):
    """Returns (trades, warnings)."""
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception as e:
        raise Mt5FormatError(
            "Ce fichier ne ressemble pas a un export MetaTrader 5 (.xlsx illisible)."
        ) from e

    ws = wb.worksheets[0]
    currency = _account_currency(ws)

    rows = list(ws.iter_rows(values_only=True))
    header_idx = None
    headers = time_idx = price_idx = None
    for i, row in enumerate(rows):
        canon = [_canonical_header(c) for c in row]
        if REQUIRED_CANONICAL <= set(canon):
            t_idx = [j for j, h in enumerate(canon) if h == "time"]
            p_idx = [j for j, h in enumerate(canon) if h == "price"]
            if len(t_idx) >= 2 and len(p_idx) >= 2:
                header_idx, headers, time_idx, price_idx = i, canon, t_idx, p_idx
                break
    if header_idx is None:
        raise Mt5FormatError(
            "Ce fichier ne ressemble pas a un rapport d'historique MetaTrader 5 "
            "(section 'Positions' introuvable)."
        )

    open_time_i, close_time_i = time_idx[0], time_idx[1]
    open_price_i, close_price_i = price_idx[0], price_idx[1]
    position_i = headers.index("position")
    symbol_i = headers.index("symbol")
    type_i = headers.index("type")
    volume_i = headers.index("volume")
    commission_i = headers.index("commission")
    swap_i = headers.index("swap")
    profit_i = headers.index("profit")

    trades = []
    for row in rows[header_idx + 1:]:
        position_id = row[position_i]
        if not isinstance(position_id, (int, float)):
            break  # end of the Positions section (next section title or blank row)

        side_raw = (row[type_i] or "").strip().lower()
        if side_raw not in ("buy", "sell"):
            continue

        symbol = clean_symbol((row[symbol_i] or "").strip())
        open_time = _parse_time(row[open_time_i])
        close_time = _parse_time(row[close_time_i])
        try:
            volume = float(str(row[volume_i]).strip())
            entry_price = float(row[open_price_i])
            exit_price = float(row[close_price_i])
        except (TypeError, ValueError):
            continue
        if not symbol or not volume or not open_time or not close_time:
            continue

        # Negated: MT5 reports these as debits (negative); the app's `commission`
        # convention is a positive cost.
        commission = -(float(row[commission_i] or 0) + float(row[swap_i] or 0))
        profit = float(row[profit_i] or 0)

        trades.append({
            "symbol": symbol,
            "side": "long" if side_raw == "buy" else "short",
            "quantity": volume,
            "entry_price": entry_price,
            "exit_price": exit_price,
            "entry_time": open_time,
            "exit_time": close_time,
            "commission": round(commission, 2),
            "pnl_native": round(profit, 2),
            "pnl_usd": round(profit, 2) if currency == "USD" else None,
            "quote_currency": currency,
            "status": "closed",
            "source": "mt5",
        })

    if not trades:
        raise Mt5FormatError("Aucune position cloturee trouvee dans ce fichier.")

    warnings = []
    if currency is None:
        warnings.append(
            "Devise du compte introuvable dans le rapport - le P&L n'est pas converti en USD."
        )
    elif currency != "USD":
        warnings.append(
            f"Compte libelle en {currency} (pas USD) - le P&L n'est pas converti en USD."
        )

    trades.sort(key=lambda t: t["entry_time"])
    return trades, warnings
